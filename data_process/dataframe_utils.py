#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@file: 
@version: 
@desc:  
@time: 2022/3/22 13:41 

@Modify Time      @Author    @Version    @Desciption
------------      -------    --------    -----------
2022/3/22 13:41   wangfc      1.0         None

用过Pandas和 openpyxl 库的同学都知道，这两个库是相互互补的。
Pandas绝对是Python中处理Excel最快、最好用的库，
但是使用 openpyxl 的一些优势是能够轻松地使用样式、条件格式等自定义电子表格。

ython3.5中第三方excel操作库- openpyxl；
其实Python第三方库有很多可以操作Excel，如：xlrd,xlwt，xlwings
甚至注明的数据分析模块Pandas也提供pandas.read_excel、pandas.DataFrame.to_excel功能。
那么openpyxl的库有哪些优缺点呢：

优势：
1、openpyxl提供对pandas的dataframe对象完美支持；
2、openpyxl支持后台静默打开excel文件；
3、它支持excel的一些sort、filter筛选、排序功能，支持丰富的单元格style（样式）设计；
4、它同时支持读取现成的excel文件&创建全新的excel文件；
5、它支持最新的xlsx格式文件，且更新频繁，操作较为简练。


缺点：
1、运算效率相对不高，当表格行项目过多时，运算相对pandas等速度较慢；
2、部分针对行或列的样式设计有一定bug，语法可能失效；
3、对sort和filter的功能虽然支持，但是需要用户手工打开excel后刷新数据方能使搜索条件生效，属于半自动；
4、不支持excel原生的自动列宽功能，实现同样效果略复杂。
"""
import io
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from sanic.request import File
from sanic.response import HTTPResponse
from openpyxl.workbook.workbook import Workbook, Worksheet
from urllib import parse
from utils.exceptions import FileIOException


def load_excel_file_to_dataframe(file: File) -> pd.DataFrame:
    try:
        # 创建一个二进制io流文件对象，这个对象才可以被load_workbook加载
        bytes_io = io.BytesIO()
        # 将文件数据传写入bytes_io对象中
        bytes_io.write(file.body)
        df = pd.read_excel(bytes_io)
        # 使用完关闭文件
        bytes_io.close()
    except Exception as e:
        raise e
    return df


def export_dataframe_to_http_response(data: pd.DataFrame, filename="demo001.xlsx") -> HTTPResponse:
    #  Pandas使用 xlsxwriter引擎保存数据,xlsxwriter模块是一个python处理Excel写入的专有模块，不支持对Excel的读取，只支持写入，功能非常强大
    # 创建ExcelWriter
    bytes_io = io.BytesIO()
    excel_writer = pd.ExcelWriter(path=bytes_io, engine='xlsxwriter')
    data.to_excel(excel_writer=excel_writer,index=False)
    excel_writer.save()
    bytes_io.seek(0)
    # 浏览器对响应头里的 Content-Disposition 里的含有中文的内容无法识别，因此出现了乱码。
    # 我们可以使用urllib对文件名进行编码, 编码格式使用utf-8即可。
    filename = parse.quote(filename, encoding="utf-8")
    http_response = HTTPResponse(body=bytes_io.getvalue(),
                                 # https://segmentfault.com/a/1190000023601065
                                 # attachment表示以文件形式下载
                                 # headers={'Content-Disposition': 'attachment;filename="{0}"'.format(filename)},
                                 headers={'Content-Disposition':
                                              "attachment;filename={0};filename*=utf-8''{0}".format(filename)},
                                 # .xls: application/vnd.ms-excel
                                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                 )
    bytes_io.close()
    return http_response


def load_excel_file_to_workbook(file: File) -> Workbook:
    """
    解析 post 一个含有 excel文件的 file 对象加载为 workbook
    """
    # 创建一个二进制io流文件对象，这个对象才可以被load_workbook加载
    bytes_io = io.BytesIO()
    # 将文件数据传写入bytes_io对象中
    bytes_io.write(file.body)
    # 这时候就可以加载到workbook对象了
    wb = load_workbook(bytes_io)
    sheet = wb.active
    # 使用完关闭文件
    bytes_io.close()

    # # 或者 先把文件读取到本地，再读取解析
    # with open(filename, 'wb') as f:  # 创建一个文件对象
    #     f.write(file.body)  # 把前端的文件内容存到文件中

    # 获取所有的sheet名称
    # sheet_names = wb.get_sheet_names()
    # 拿到第一个sheet，根据实际情况来看
    # sheet_name = sheet_names[0]
    # title_list = sheet.values  # 第一行的行头或者列名
    # sheet = wb[sheet_name]  # 创建工作簿对象
    # max_row = sheet.max_row  # 获取最大行数（自动识别有数据的行，如果数据行是连续的）
    # max_col = sheet.max_col  # 获取最大列数（自动识别有数据的列，也要数据连续才行）

    # 下面遍历每行，拿到数据，第一行为行头，略过
    # data = []
    # for row in sheet.iter_rows(min_row=2, max_col=max_col, max_row=sheet.max_row, values_only=True):
    #     data.append(row)
    return wb


def openpyxl_sheet_to_dataframe(filename):
    # 创建workbook对象
    workbook = load_workbook(filename=filename)
    # 默认的工作簿
    sheet = workbook.active
    values = sheet.values
    df = pd.DataFrame(values)
    return df
